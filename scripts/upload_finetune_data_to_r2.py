#!/usr/bin/env python3
"""
Comprehensive script to upload all CollegeAdvisory Finetune Data to R2 bucket.
Analyzes PDF and XLSX files, generates detailed reports, and cleans up local storage.
"""

import sys
import os
import json
import shutil
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Tuple
import logging

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from college_advisor_data.storage.r2_storage import R2StorageClient

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('finetune_data_upload.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class FinetuneDataUploader:
    """Handles uploading, analyzing, and managing finetune data."""
    
    def __init__(self, source_dir: str = "CollegeAdvisory Finetune Data"):
        self.source_dir = Path(source_dir)
        self.r2_client = None
        self.stats = {
            "total_files": 0,
            "pdf_files": 0,
            "xlsx_files": 0,
            "other_files": 0,
            "total_size_mb": 0,
            "uploaded_files": 0,
            "failed_uploads": 0,
            "analyzed_files": 0,
            "deleted_files": 0,
            "start_time": datetime.now().isoformat(),
            "end_time": None
        }
        self.file_analysis = []
        self.upload_results = []
        
    def initialize_r2_client(self):
        """Initialize R2 storage client."""
        try:
            self.r2_client = R2StorageClient()
            logger.info("✅ R2 client initialized successfully")
            
            # Ensure bucket exists
            self.r2_client.create_bucket()
            logger.info(f"✅ R2 bucket '{self.r2_client.bucket_name}' ready")
            return True
        except Exception as e:
            logger.error(f"❌ Failed to initialize R2 client: {e}")
            return False
    
    def scan_directory(self) -> List[Path]:
        """Scan source directory and categorize files."""
        if not self.source_dir.exists():
            logger.error(f"❌ Directory not found: {self.source_dir}")
            return []
        
        logger.info(f"📂 Scanning directory: {self.source_dir}")
        files = list(self.source_dir.glob("*"))
        files = [f for f in files if f.is_file()]
        
        self.stats["total_files"] = len(files)
        
        for file in files:
            ext = file.suffix.lower()
            size_mb = file.stat().st_size / (1024 * 1024)
            self.stats["total_size_mb"] += size_mb
            
            if ext == ".pdf":
                self.stats["pdf_files"] += 1
            elif ext in [".xlsx", ".xls"]:
                self.stats["xlsx_files"] += 1
            else:
                self.stats["other_files"] += 1
        
        logger.info(f"📊 Found {self.stats['total_files']} files:")
        logger.info(f"   - PDF files: {self.stats['pdf_files']}")
        logger.info(f"   - XLSX files: {self.stats['xlsx_files']}")
        logger.info(f"   - Other files: {self.stats['other_files']}")
        logger.info(f"   - Total size: {self.stats['total_size_mb']:.2f} MB")
        
        return sorted(files)
    
    def analyze_pdf_file(self, file_path: Path) -> Dict[str, Any]:
        """Analyze PDF file structure and content."""
        try:
            import PyPDF2
            
            analysis = {
                "filename": file_path.name,
                "type": "PDF",
                "size_mb": file_path.stat().st_size / (1024 * 1024),
                "pages": 0,
                "text_sample": "",
                "metadata": {},
                "analysis_status": "success"
            }
            
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                analysis["pages"] = len(pdf_reader.pages)
                
                # Extract metadata
                if pdf_reader.metadata:
                    analysis["metadata"] = {
                        k: str(v) for k, v in pdf_reader.metadata.items()
                    }
                
                # Extract text sample from first page
                if len(pdf_reader.pages) > 0:
                    first_page = pdf_reader.pages[0]
                    text = first_page.extract_text()
                    analysis["text_sample"] = text[:500] if text else ""
                    
                    # Detect if it's a Common Data Set
                    if "common data set" in text.lower():
                        analysis["document_type"] = "Common Data Set"
                    
            return analysis
            
        except Exception as e:
            logger.warning(f"⚠️  Could not analyze PDF {file_path.name}: {e}")
            return {
                "filename": file_path.name,
                "type": "PDF",
                "size_mb": file_path.stat().st_size / (1024 * 1024),
                "analysis_status": "failed",
                "error": str(e)
            }
    
    def analyze_xlsx_file(self, file_path: Path) -> Dict[str, Any]:
        """Analyze XLSX file structure and content."""
        try:
            import openpyxl
            
            analysis = {
                "filename": file_path.name,
                "type": "XLSX",
                "size_mb": file_path.stat().st_size / (1024 * 1024),
                "sheets": [],
                "total_rows": 0,
                "total_columns": 0,
                "analysis_status": "success"
            }
            
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                max_row = sheet.max_row or 0
                max_col = sheet.max_column or 0
                
                sheet_info = {
                    "name": sheet_name,
                    "rows": max_row,
                    "columns": max_col
                }
                
                # Sample first few cells
                sample_data = []
                for row in sheet.iter_rows(min_row=1, max_row=min(5, max_row), values_only=True):
                    sample_data.append([str(cell)[:50] if cell else "" for cell in row[:5]])
                
                sheet_info["sample_data"] = sample_data
                analysis["sheets"].append(sheet_info)
                analysis["total_rows"] += max_row
                analysis["total_columns"] = max(analysis["total_columns"], max_col)
            
            wb.close()
            return analysis
            
        except Exception as e:
            logger.warning(f"⚠️  Could not analyze XLSX {file_path.name}: {e}")
            return {
                "filename": file_path.name,
                "type": "XLSX",
                "size_mb": file_path.stat().st_size / (1024 * 1024),
                "analysis_status": "failed",
                "error": str(e)
            }
    
    def analyze_file(self, file_path: Path) -> Dict[str, Any]:
        """Analyze file based on its type."""
        ext = file_path.suffix.lower()
        
        if ext == ".pdf":
            return self.analyze_pdf_file(file_path)
        elif ext in [".xlsx", ".xls"]:
            return self.analyze_xlsx_file(file_path)
        else:
            return {
                "filename": file_path.name,
                "type": ext.upper().replace(".", ""),
                "size_mb": file_path.stat().st_size / (1024 * 1024),
                "analysis_status": "skipped"
            }
    
    def upload_file_to_r2(self, file_path: Path) -> Tuple[bool, str]:
        """Upload a single file to R2 bucket."""
        try:
            # Create R2 key with organized structure
            r2_key = f"finetune_data/common_data_sets/{file_path.name}"
            
            # Prepare metadata
            metadata = {
                "source": "CollegeAdvisory Finetune Data",
                "upload_date": datetime.now().isoformat(),
                "file_type": file_path.suffix.lower(),
                "original_size": str(file_path.stat().st_size)
            }
            
            # Upload file
            success = self.r2_client.upload_file(
                file_path=file_path,
                object_key=r2_key,
                metadata=metadata
            )
            
            if success:
                self.stats["uploaded_files"] += 1
                return True, r2_key
            else:
                self.stats["failed_uploads"] += 1
                return False, ""
                
        except Exception as e:
            logger.error(f"❌ Failed to upload {file_path.name}: {e}")
            self.stats["failed_uploads"] += 1
            return False, str(e)
    
    def process_all_files(self, files: List[Path]):
        """Process all files: analyze and upload."""
        logger.info(f"\n{'='*80}")
        logger.info("PROCESSING FILES")
        logger.info(f"{'='*80}\n")
        
        for idx, file_path in enumerate(files, 1):
            logger.info(f"[{idx}/{len(files)}] Processing: {file_path.name}")
            
            # Analyze file
            logger.info("   📊 Analyzing file...")
            analysis = self.analyze_file(file_path)
            self.file_analysis.append(analysis)
            
            if analysis["analysis_status"] == "success":
                self.stats["analyzed_files"] += 1
                logger.info(f"   ✅ Analysis complete")
            
            # Upload to R2
            logger.info("   ☁️  Uploading to R2...")
            success, r2_key_or_error = self.upload_file_to_r2(file_path)
            
            upload_result = {
                "filename": file_path.name,
                "success": success,
                "r2_key": r2_key_or_error if success else None,
                "error": r2_key_or_error if not success else None,
                "timestamp": datetime.now().isoformat()
            }
            self.upload_results.append(upload_result)
            
            if success:
                logger.info(f"   ✅ Uploaded to: {r2_key_or_error}")
            else:
                logger.error(f"   ❌ Upload failed: {r2_key_or_error}")
            
            logger.info("")

    def verify_uploads(self) -> bool:
        """Verify all files were uploaded successfully to R2."""
        logger.info(f"\n{'='*80}")
        logger.info("VERIFYING R2 UPLOADS")
        logger.info(f"{'='*80}\n")

        try:
            # List all objects in the finetune_data prefix
            response = self.r2_client.client.list_objects_v2(
                Bucket=self.r2_client.bucket_name,
                Prefix="finetune_data/common_data_sets/"
            )

            if 'Contents' not in response:
                logger.error("❌ No files found in R2 bucket!")
                return False

            r2_files = {obj['Key'].split('/')[-1]: obj for obj in response['Contents']}
            logger.info(f"✅ Found {len(r2_files)} files in R2 bucket")

            # Verify each uploaded file
            missing_files = []
            for result in self.upload_results:
                if result["success"]:
                    filename = result["filename"]
                    if filename not in r2_files:
                        missing_files.append(filename)
                        logger.error(f"❌ Missing in R2: {filename}")

            if missing_files:
                logger.error(f"❌ {len(missing_files)} files missing from R2!")
                return False

            logger.info("✅ All uploaded files verified in R2 bucket")
            return True

        except Exception as e:
            logger.error(f"❌ Verification failed: {e}")
            return False

    def delete_local_files(self, files: List[Path]):
        """Delete local files after successful upload and verification."""
        logger.info(f"\n{'='*80}")
        logger.info("DELETING LOCAL FILES")
        logger.info(f"{'='*80}\n")

        deleted_count = 0
        failed_count = 0

        for file_path in files:
            # Only delete if upload was successful
            upload_result = next(
                (r for r in self.upload_results if r["filename"] == file_path.name),
                None
            )

            if upload_result and upload_result["success"]:
                try:
                    file_path.unlink()
                    deleted_count += 1
                    logger.info(f"🗑️  Deleted: {file_path.name}")
                except Exception as e:
                    failed_count += 1
                    logger.error(f"❌ Failed to delete {file_path.name}: {e}")
            else:
                logger.warning(f"⚠️  Skipped deletion (upload failed): {file_path.name}")

        self.stats["deleted_files"] = deleted_count
        logger.info(f"\n✅ Deleted {deleted_count} local files")

        if failed_count > 0:
            logger.warning(f"⚠️  Failed to delete {failed_count} files")

        # Try to remove directory if empty
        try:
            remaining_files = list(self.source_dir.glob("*"))
            if not remaining_files:
                self.source_dir.rmdir()
                logger.info(f"🗑️  Removed empty directory: {self.source_dir}")
        except Exception as e:
            logger.info(f"ℹ️  Directory not removed: {e}")

    def generate_report(self) -> Dict[str, Any]:
        """Generate comprehensive report on data processing."""
        logger.info(f"\n{'='*80}")
        logger.info("GENERATING COMPREHENSIVE REPORT")
        logger.info(f"{'='*80}\n")

        self.stats["end_time"] = datetime.now().isoformat()

        # Categorize files by university
        universities = {}
        for analysis in self.file_analysis:
            filename = analysis["filename"]
            # Extract university name (before first dash or underscore)
            parts = filename.replace("-", "_").split("_")
            if len(parts) > 0:
                uni_name = parts[0]
                if uni_name not in universities:
                    universities[uni_name] = []
                universities[uni_name].append(filename)

        # Analyze Common Data Set structure
        cds_insights = {
            "total_cds_documents": 0,
            "avg_pages_per_pdf": 0,
            "total_pages": 0,
            "universities_covered": len(universities),
            "file_formats": {
                "pdf": self.stats["pdf_files"],
                "xlsx": self.stats["xlsx_files"],
                "other": self.stats["other_files"]
            }
        }

        total_pages = 0
        pdf_count = 0
        for analysis in self.file_analysis:
            if analysis.get("document_type") == "Common Data Set":
                cds_insights["total_cds_documents"] += 1
            if analysis["type"] == "PDF" and "pages" in analysis:
                total_pages += analysis["pages"]
                pdf_count += 1

        if pdf_count > 0:
            cds_insights["avg_pages_per_pdf"] = total_pages / pdf_count
            cds_insights["total_pages"] = total_pages

        report = {
            "summary": self.stats,
            "cds_insights": cds_insights,
            "universities": universities,
            "file_analysis": self.file_analysis,
            "upload_results": self.upload_results,
            "recommendations": self._generate_recommendations()
        }

        # Save report to file
        report_path = Path("finetune_data_upload_report.json")
        with open(report_path, 'w') as f:
            json.dump(report, f, indent=2)

        logger.info(f"✅ Report saved to: {report_path}")

        # Print summary
        self._print_summary(report)

        return report

    def _generate_recommendations(self) -> List[str]:
        """Generate recommendations for data integration."""
        recommendations = [
            "DATA INTEGRATION RECOMMENDATIONS:",
            "",
            "1. COMMON DATA SET STRUCTURE:",
            "   - Files contain standardized university data (CDS format)",
            "   - Sections typically include: General Info, Enrollment, Admissions, Financial Aid, etc.",
            "   - Recommend parsing PDFs to extract structured data for each section",
            "",
            "2. PROCESSING PIPELINE:",
            "   - Use PyPDF2 or pdfplumber to extract text from PDFs",
            "   - Use openpyxl or pandas to process XLSX files",
            "   - Create unified schema mapping CDS sections to database fields",
            "   - Implement validation to ensure data quality",
            "",
            "3. FINETUNING DATA PREPARATION:",
            "   - Extract Q&A pairs from CDS documents (e.g., 'What is the acceptance rate at MIT?')",
            "   - Create instruction-following examples for admissions data",
            "   - Generate comparison queries across universities",
            "   - Include context about data sources and years",
            "",
            "4. DATA QUALITY CHECKS:",
            "   - Verify all universities have complete CDS data",
            "   - Check for missing sections or incomplete information",
            "   - Validate numerical data (acceptance rates, enrollment, etc.)",
            "   - Ensure year consistency across datasets",
            "",
            "5. STORAGE AND RETRIEVAL:",
            "   - Files now stored in R2 under 'finetune_data/common_data_sets/'",
            "   - Implement caching strategy for frequently accessed data",
            "   - Create index for quick university lookup",
            "   - Consider creating processed/structured versions in separate R2 prefix",
            "",
            "6. NEXT STEPS:",
            "   - Implement CDS parser to extract structured data",
            "   - Create training data generator for finetuning",
            "   - Integrate with existing RAG pipeline",
            "   - Set up automated processing for new CDS files"
        ]

        return recommendations

    def _print_summary(self, report: Dict[str, Any]):
        """Print summary of processing results."""
        logger.info(f"\n{'='*80}")
        logger.info("PROCESSING SUMMARY")
        logger.info(f"{'='*80}\n")

        stats = report["summary"]
        insights = report["cds_insights"]

        logger.info(f"📊 FILES PROCESSED:")
        logger.info(f"   Total files: {stats['total_files']}")
        logger.info(f"   PDF files: {stats['pdf_files']}")
        logger.info(f"   XLSX files: {stats['xlsx_files']}")
        logger.info(f"   Total size: {stats['total_size_mb']:.2f} MB")
        logger.info("")

        logger.info(f"☁️  R2 UPLOADS:")
        logger.info(f"   Successful: {stats['uploaded_files']}")
        logger.info(f"   Failed: {stats['failed_uploads']}")
        logger.info("")

        logger.info(f"📖 ANALYSIS:")
        logger.info(f"   Files analyzed: {stats['analyzed_files']}")
        logger.info(f"   CDS documents: {insights['total_cds_documents']}")
        logger.info(f"   Total pages: {insights['total_pages']}")
        logger.info(f"   Avg pages/PDF: {insights['avg_pages_per_pdf']:.1f}")
        logger.info(f"   Universities: {insights['universities_covered']}")
        logger.info("")

        logger.info(f"🗑️  CLEANUP:")
        logger.info(f"   Files deleted: {stats['deleted_files']}")
        logger.info("")

        logger.info(f"⏱️  TIMING:")
        logger.info(f"   Start: {stats['start_time']}")
        logger.info(f"   End: {stats['end_time']}")
        logger.info("")

        # Print recommendations
        for rec in report["recommendations"]:
            logger.info(rec)


def main():
    """Main execution function."""
    logger.info(f"\n{'='*80}")
    logger.info("FINETUNE DATA UPLOAD TO R2 BUCKET")
    logger.info(f"{'='*80}\n")

    uploader = FinetuneDataUploader()

    # Step 1: Initialize R2 client
    logger.info("Step 1: Initializing R2 client...")
    if not uploader.initialize_r2_client():
        logger.error("❌ Failed to initialize R2 client. Exiting.")
        return False

    # Step 2: Scan directory
    logger.info("\nStep 2: Scanning source directory...")
    files = uploader.scan_directory()
    if not files:
        logger.error("❌ No files found. Exiting.")
        return False

    # Step 3: Process all files (analyze + upload)
    logger.info("\nStep 3: Processing files (analyze + upload)...")
    uploader.process_all_files(files)

    # Step 4: Verify uploads
    logger.info("\nStep 4: Verifying uploads...")
    verification_success = uploader.verify_uploads()

    if not verification_success:
        logger.error("❌ Upload verification failed. NOT deleting local files.")
        uploader.generate_report()
        return False

    # Step 5: Delete local files
    logger.info("\nStep 5: Deleting local files...")
    uploader.delete_local_files(files)

    # Step 6: Generate comprehensive report
    logger.info("\nStep 6: Generating report...")
    uploader.generate_report()

    logger.info(f"\n{'='*80}")
    logger.info("✅ ALL OPERATIONS COMPLETED SUCCESSFULLY")
    logger.info(f"{'='*80}\n")

    return True


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)


