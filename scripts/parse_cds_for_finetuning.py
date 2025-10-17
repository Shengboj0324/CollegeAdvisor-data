#!/usr/bin/env python3
"""
Parse Common Data Set documents from R2 and generate finetuning data.
This script downloads CDS files, extracts structured data, and creates Q&A pairs.
"""

import sys
import json
import re
from pathlib import Path
from typing import Dict, List, Any, Tuple
import logging

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from college_advisor_data.storage.r2_storage import R2StorageClient

logging.basicConfig(level=logging.INFO, format='%(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class CDSParser:
    """Parse Common Data Set documents and extract structured data."""
    
    def __init__(self):
        self.r2_client = R2StorageClient()
        self.parsed_data = []
        self.qa_pairs = []
        
    def download_cds_files(self, output_dir: Path = Path("data/cds_downloads")):
        """Download all CDS files from R2."""
        output_dir.mkdir(parents=True, exist_ok=True)
        
        logger.info("Downloading CDS files from R2...")
        
        # List all files in R2
        response = self.r2_client.client.list_objects_v2(
            Bucket=self.r2_client.bucket_name,
            Prefix="finetune_data/common_data_sets/"
        )
        
        if 'Contents' not in response:
            logger.error("No files found in R2!")
            return []
        
        downloaded_files = []
        for obj in response['Contents']:
            key = obj['Key']
            filename = key.split('/')[-1]
            
            # Skip non-data files
            if filename.startswith('.') or filename.startswith('_$'):
                continue
            
            local_path = output_dir / filename
            
            if self.r2_client.download_file(key, local_path):
                downloaded_files.append(local_path)
                logger.info(f"Downloaded: {filename}")
        
        logger.info(f"Downloaded {len(downloaded_files)} files")
        return downloaded_files
    
    def extract_pdf_text(self, pdf_path: Path) -> str:
        """Extract text from PDF file."""
        try:
            import PyPDF2
            
            text = ""
            with open(pdf_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
            
            return text
        except Exception as e:
            logger.error(f"Failed to extract text from {pdf_path.name}: {e}")
            return ""
    
    def extract_xlsx_data(self, xlsx_path: Path) -> Dict[str, Any]:
        """Extract data from XLSX file."""
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            data = {}
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_data = []
                
                for row in sheet.iter_rows(values_only=True):
                    sheet_data.append(list(row))
                
                data[sheet_name] = sheet_data
            
            wb.close()
            return data
        except Exception as e:
            logger.error(f"Failed to extract data from {xlsx_path.name}: {e}")
            return {}
    
    def parse_university_name(self, filename: str) -> str:
        """Extract university name from filename."""
        # Remove file extension
        name = filename.rsplit('.', 1)[0]
        
        # Common patterns
        if '-' in name:
            # Format: "University Name-CDS_2024-2025.pdf"
            parts = name.split('-')
            if len(parts) > 0:
                return parts[0].strip()
        
        if '_' in name:
            # Format: "University_Name_CDS_2024_2025.pdf"
            parts = name.split('_')
            # Find where CDS or year starts
            for i, part in enumerate(parts):
                if 'cds' in part.lower() or part.isdigit():
                    return ' '.join(parts[:i]).strip()
        
        return name
    
    def extract_acceptance_rate(self, text: str) -> float:
        """Extract acceptance rate from CDS text."""
        # Look for patterns like "X% acceptance rate" or "admitted X%"
        patterns = [
            r'acceptance rate[:\s]+(\d+\.?\d*)%',
            r'admitted[:\s]+(\d+\.?\d*)%',
            r'(\d+\.?\d*)%\s+acceptance',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return float(match.group(1))
        
        return None
    
    def extract_enrollment(self, text: str) -> int:
        """Extract total enrollment from CDS text."""
        # Look for total enrollment numbers
        patterns = [
            r'total enrollment[:\s]+(\d+,?\d*)',
            r'enrollment[:\s]+(\d+,?\d*)\s+students',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                num_str = match.group(1).replace(',', '')
                return int(num_str)
        
        return None
    
    def generate_qa_pairs_from_cds(self, university: str, data: Dict[str, Any]) -> List[Dict[str, str]]:
        """Generate Q&A pairs from parsed CDS data."""
        qa_pairs = []
        
        # Acceptance rate Q&A
        if data.get('acceptance_rate'):
            qa_pairs.append({
                "instruction": f"What is the acceptance rate at {university}?",
                "input": "",
                "output": f"According to the Common Data Set, {university} has an acceptance rate of {data['acceptance_rate']}%."
            })
        
        # Enrollment Q&A
        if data.get('enrollment'):
            qa_pairs.append({
                "instruction": f"How many students are enrolled at {university}?",
                "input": "",
                "output": f"Based on the Common Data Set, {university} has a total enrollment of {data['enrollment']:,} students."
            })
        
        # Tuition Q&A (if available)
        if data.get('tuition'):
            qa_pairs.append({
                "instruction": f"What is the tuition at {university}?",
                "input": "",
                "output": f"According to the Common Data Set, the tuition at {university} is ${data['tuition']:,} per year."
            })
        
        # General info Q&A
        qa_pairs.append({
            "instruction": f"Tell me about {university}",
            "input": "",
            "output": f"{university} is a university with {data.get('enrollment', 'N/A')} students and an acceptance rate of {data.get('acceptance_rate', 'N/A')}%. For more detailed information, please refer to the Common Data Set."
        })
        
        return qa_pairs
    
    def process_all_files(self, files: List[Path]):
        """Process all downloaded CDS files."""
        logger.info(f"\nProcessing {len(files)} CDS files...")
        
        for file_path in files:
            logger.info(f"Processing: {file_path.name}")
            
            university = self.parse_university_name(file_path.name)
            
            data = {
                "university": university,
                "filename": file_path.name,
                "source": "Common Data Set"
            }
            
            # Extract data based on file type
            if file_path.suffix.lower() == '.pdf':
                text = self.extract_pdf_text(file_path)
                if text:
                    data['acceptance_rate'] = self.extract_acceptance_rate(text)
                    data['enrollment'] = self.extract_enrollment(text)
            
            elif file_path.suffix.lower() in ['.xlsx', '.xls']:
                xlsx_data = self.extract_xlsx_data(file_path)
                # TODO: Extract specific fields from XLSX sheets
                data['xlsx_sheets'] = list(xlsx_data.keys())
            
            self.parsed_data.append(data)
            
            # Generate Q&A pairs
            qa_pairs = self.generate_qa_pairs_from_cds(university, data)
            self.qa_pairs.extend(qa_pairs)
        
        logger.info(f"Processed {len(self.parsed_data)} files")
        logger.info(f"Generated {len(self.qa_pairs)} Q&A pairs")
    
    def save_training_data(self, output_path: Path = Path("data/training/cds_training_data.json")):
        """Save Q&A pairs in training format."""
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w') as f:
            json.dump(self.qa_pairs, f, indent=2)
        
        logger.info(f"Saved training data to: {output_path}")
        
        # Also save in Alpaca format
        alpaca_path = output_path.parent / "cds_training_data_alpaca.json"
        with open(alpaca_path, 'w') as f:
            json.dump(self.qa_pairs, f, indent=2)
        
        logger.info(f"Saved Alpaca format to: {alpaca_path}")
    
    def save_parsed_data(self, output_path: Path = Path("data/processed/cds_parsed_data.json")):
        """Save parsed structured data."""
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w') as f:
            json.dump(self.parsed_data, f, indent=2)
        
        logger.info(f"Saved parsed data to: {output_path}")
    
    def generate_statistics(self):
        """Generate statistics about the parsed data."""
        stats = {
            "total_universities": len(self.parsed_data),
            "total_qa_pairs": len(self.qa_pairs),
            "universities_with_acceptance_rate": sum(1 for d in self.parsed_data if d.get('acceptance_rate')),
            "universities_with_enrollment": sum(1 for d in self.parsed_data if d.get('enrollment')),
            "avg_acceptance_rate": None,
            "avg_enrollment": None
        }
        
        # Calculate averages
        acceptance_rates = [d['acceptance_rate'] for d in self.parsed_data if d.get('acceptance_rate')]
        if acceptance_rates:
            stats['avg_acceptance_rate'] = sum(acceptance_rates) / len(acceptance_rates)
        
        enrollments = [d['enrollment'] for d in self.parsed_data if d.get('enrollment')]
        if enrollments:
            stats['avg_enrollment'] = sum(enrollments) / len(enrollments)
        
        return stats


def main():
    """Main execution function."""
    logger.info("="*80)
    logger.info("CDS PARSER FOR FINETUNING DATA GENERATION")
    logger.info("="*80)
    
    parser = CDSParser()
    
    # Step 1: Download files from R2
    logger.info("\nStep 1: Downloading CDS files from R2...")
    files = parser.download_cds_files()
    
    if not files:
        logger.error("No files downloaded. Exiting.")
        return False
    
    # Step 2: Process all files
    logger.info("\nStep 2: Processing CDS files...")
    parser.process_all_files(files)
    
    # Step 3: Save training data
    logger.info("\nStep 3: Saving training data...")
    parser.save_training_data()
    
    # Step 4: Save parsed data
    logger.info("\nStep 4: Saving parsed data...")
    parser.save_parsed_data()
    
    # Step 5: Generate statistics
    logger.info("\nStep 5: Generating statistics...")
    stats = parser.generate_statistics()
    
    logger.info("\n" + "="*80)
    logger.info("PROCESSING COMPLETE")
    logger.info("="*80)
    logger.info(f"Total Universities: {stats['total_universities']}")
    logger.info(f"Total Q&A Pairs: {stats['total_qa_pairs']}")
    logger.info(f"Universities with Acceptance Rate: {stats['universities_with_acceptance_rate']}")
    logger.info(f"Universities with Enrollment: {stats['universities_with_enrollment']}")
    if stats['avg_acceptance_rate']:
        logger.info(f"Average Acceptance Rate: {stats['avg_acceptance_rate']:.2f}%")
    if stats['avg_enrollment']:
        logger.info(f"Average Enrollment: {stats['avg_enrollment']:,.0f}")
    logger.info("="*80)
    
    return True


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)

